"""
工事台帳 サーバー v3
- LINEなし・Webアプリ専用
- JWT認証（ID/パスワード）
- Claude APIで画像読み取り
- PCエージェント向けキューAPI
"""
import os, json, base64, re, hashlib, secrets, io, tempfile
from pathlib import Path
from datetime import datetime, timedelta

import anthropic
from fastapi import FastAPI, Request, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import jwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Google Drive
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

app = FastAPI()

# CORS（Webアプリからのアクセスを許可）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 本番時はフロントのURLに絞る
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# 設定
# =========================================================
CLAUDE_KEY  = os.environ["ANTHROPIC_API_KEY"]
JWT_SECRET  = os.environ.get("JWT_SECRET", secrets.token_hex(32))
JWT_EXPIRE  = 60 * 24 * 7  # 7日間（分）
DATA_FILE   = Path("data/store.json")
DATA_FILE.parent.mkdir(exist_ok=True)

CATEGORIES  = ["材料費", "人件費", "外注費", "経費"]
security    = HTTPBearer()
CAT_COLORS  = {"材料費": "1565C0", "人件費": "2E7D32", "外注費": "6A1B9A", "経費": "E65100"}
GDRIVE_FOLDER_ID = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")

def get_drive_service():
    """Google Drive サービスを取得"""
    if not GDRIVE_AVAILABLE or not GDRIVE_FOLDER_ID:
        return None
    creds_b64 = os.environ.get("GOOGLE_CREDENTIALS_B64", "")
    if not creds_b64:
        return None
    try:
        # 改行・スペース・不要文字を除去してからデコード
        creds_b64_clean = re.sub(r'[\s]', '', creds_b64)
        # base64パディング補正
        padding = 4 - len(creds_b64_clean) % 4
        if padding != 4:
            creds_b64_clean += '=' * padding
        creds_json = base64.b64decode(creds_b64_clean).decode("utf-8")
        # JSONの余分な文字を除去
        creds_json = creds_json.strip()
        # 複数のJSONが連結されている場合は最初のJSONオブジェクトだけ使う
        brace_count = 0
        end_idx = 0
        for i, ch in enumerate(creds_json):
            if ch == '{':
                brace_count += 1
            elif ch == '}':
                brace_count -= 1
                if brace_count == 0:
                    end_idx = i + 1
                    break
        creds_json = creds_json[:end_idx]
        creds_info = json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(
            creds_info,
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        print(f"Google Drive接続エラー: {e}")
        return None

# =========================================================
# データ管理
# =========================================================
def load() -> dict:
    if DATA_FILE.exists():
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    return {"users": [], "projects": [], "queue": []}

def save(data: dict):
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def active_projects(data: dict) -> list:
    return [p for p in data["projects"] if not p.get("done")]

# =========================================================
# 認証
# =========================================================
def create_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.utcnow() + timedelta(minutes=JWT_EXPIRE),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=["HS256"])
        return payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "トークンが期限切れです")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "認証エラーです")

# =========================================================
# 認証エンドポイント
# =========================================================
@app.post("/api/auth/register")
async def register(request: Request):
    body = await request.json()
    username = body.get("username", "").strip()
    password = body.get("password", "")
    if not username or not password:
        raise HTTPException(400, "ユーザー名とパスワードを入力してください")
    data = load()
    if any(u["username"] == username for u in data["users"]):
        raise HTTPException(409, "このユーザー名はすでに使われています")
    user = {"id": f"U{int(datetime.now().timestamp())}", "username": username, "password": hash_password(password)}
    data["users"].append(user)
    save(data)
    return {"token": create_token(user["id"]), "username": username}

@app.post("/api/auth/login")
async def login(request: Request):
    body = await request.json()
    username = body.get("username", "").strip()
    password = body.get("password", "")
    data = load()
    user = next((u for u in data["users"] if u["username"] == username and u["password"] == hash_password(password)), None)
    if not user:
        raise HTTPException(401, "ユーザー名またはパスワードが違います")
    return {"token": create_token(user["id"]), "username": username}

# =========================================================
# 工事管理
# =========================================================
@app.get("/api/projects")
async def get_projects(user_id: str = Depends(verify_token)):
    data = load()
    return {"projects": [p for p in data["projects"] if p.get("owner") == user_id]}

@app.post("/api/projects")
async def create_project(request: Request, user_id: str = Depends(verify_token)):
    body    = await request.json()
    name    = body.get("name", "").strip()
    if not name:
        raise HTTPException(400, "工事名を入力してください")
    data    = load()
    num     = body.get("num") or f"{datetime.now().year}-{len(data['projects'])+1:03d}"
    project = {
        "id":              f"P{int(datetime.now().timestamp())}",
        "name":            name,
        "num":             num,
        "start":           body.get("start", datetime.now().strftime("%Y-%m-%d")),
        "person":          body.get("person", ""),
        "location":        body.get("location", ""),
        "contract_amount":     body.get("contract_amount", ""),
        "contract_amount_ex":  body.get("contract_amount_ex", ""),
        "contract_amount_tax": body.get("contract_amount_tax", ""),
        "orderer":             body.get("orderer", ""),           # 注文者
        "jv_type":         body.get("jv_type", "元請"),       # 元請/下請
        "engineer_name":   body.get("engineer_name", ""),     # 配置技術者氏名
        "engineer_chief":  body.get("engineer_chief", ""),    # 主任技術者
        "engineer_super":  body.get("engineer_super", ""),    # 監理技術者
        "has_pc":          body.get("has_pc", False),         # PC
        "has_surface":     body.get("has_surface", False),    # 法面処理
        "has_steel":       body.get("has_steel", False),      # 鋼橋上部
        "done":            False,
        "owner":           user_id,
    }
    data["projects"].append(project)
    save(data)
    return {"project": project}

@app.patch("/api/projects/{project_id}/done")
async def complete_project(project_id: str, request: Request, user_id: str = Depends(verify_token)):
    body = await request.json() if request.headers.get("content-type") == "application/json" else {}
    completion_date = body.get("completion_date", "") if body else ""
    data = load()
    for p in data["projects"]:
        if p["id"] == project_id and p.get("owner") == user_id:
            p["done"] = True
            if completion_date:
                p["completion_date"] = completion_date
            save(data)
            return {"ok": True}
    raise HTTPException(404, "工事が見つかりません")

@app.patch("/api/projects/{project_id}")
async def update_project(project_id: str, request: Request, user_id: str = Depends(verify_token)):
    """工事経歴書用の追加情報を更新する"""
    body = await request.json()
    data = load()
    for p in data["projects"]:
        if p["id"] == project_id and p.get("owner") == user_id:
            updatable = ["orderer","jv_type","engineer_name","engineer_chief","engineer_super",
                         "has_pc","has_surface","has_steel","location","contract_amount",
                         "contract_amount_ex","contract_amount_tax","person","num","name","start"]
            for key in updatable:
                if key in body:
                    p[key] = body[key]
            save(data)
            return {"ok": True, "project": p}
    raise HTTPException(404, "工事が見つかりません")

@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: str, user_id: str = Depends(verify_token)):
    data = load()
    before = len(data["projects"])
    data["projects"] = [p for p in data["projects"] if not (p["id"] == project_id and p.get("owner") == user_id)]
    if len(data["projects"]) == before:
        raise HTTPException(404, "工事が見つかりません")
    data["queue"] = [r for r in data["queue"] if r["project_id"] != project_id]
    save(data)
    return {"ok": True}

# =========================================================
# AI画像読み取り → キューへ追加
# =========================================================
@app.post("/api/records")
async def create_record(request: Request, user_id: str = Depends(verify_token)):
    body       = await request.json()
    project_id = body.get("project_id")
    category   = body.get("category")
    image_b64  = body.get("image_b64")  # base64エンコード済み画像

    if not all([project_id, category, image_b64]):
        raise HTTPException(400, "project_id・category・image_b64 は必須です")
    if category not in CATEGORIES:
        raise HTTPException(400, f"費目は {CATEGORIES} のいずれかを指定してください")

    data    = load()
    project = next((p for p in data["projects"] if p["id"] == project_id), None)
    if not project:
        raise HTTPException(404, "工事が見つかりません")

    # Claude APIで読み取り
    try:
        result = await ai_read(base64.b64decode(image_b64), category)
    except ValueError:
        raise HTTPException(422, "画像を読み取れませんでした。明るく・文字がはっきり写った写真で再度お試しください。")
    except Exception:
        raise HTTPException(500, "AI読み取り中にエラーが発生しました。時間をおいて再度お試しください。")

    # キューに追加
    record = {
        "id":           f"R{int(datetime.now().timestamp())}",
        "project_id":   project_id,
        "project_name": project["name"],
        "project_num":  project["num"],
        "category":     category,
        "ai_result":    result,
        "queued_at":    datetime.now().isoformat(),
        "done":         False,
    }
    data["queue"].append(record)
    save(data)
    return {"record": record, "ai_result": result}

# =========================================================
# Claude API 画像読み取り（高精度版：Opus + 2段階読み取り）
# =========================================================
async def ai_read(image_bytes: bytes, category: str, media_type: str = "image/jpeg") -> dict:
    client   = anthropic.Anthropic(api_key=CLAUDE_KEY)
    b64      = base64.standard_b64encode(image_bytes).decode()
    is_pdf   = media_type == "application/pdf"

    cat_hint = {
        "材料費": "建築・土木工事の材料購入レシート・納品書（セメント・木材・鉄筋・砂利・塗料・ボルト・釘など）",
        "人件費": "作業員・職人への賃金支払い領収書・作業日報（氏名・作業内容・日数・時間・単価など）",
        "外注費": "下請け業者・専門業者への支払い領収書（業者名・作業内容・工事名・金額）",
        "経費":   "工事に関わる交通費・高速代・駐車場代・消耗品・工具代・ガソリン代などの領収書・レシート",
        "不明":   "建築土木工事に関連する領収書・レシート・納品書・請求書",
    }.get(category, "建築土木工事に関連する領収書・レシート・納品書")

    system = """あなたは建築土木業の経理担当として20年以上の経験を持つ専門家です。
日本の手書き領収書・印刷レシート・納品書・請求書の読み取りに特化しています。

【絶対ルール】
- JSONのみを返す。前置き・説明・```などのコードブロック記号は一切不要
- 読み取れない項目はnullとし、絶対に推測で埋めない
- 金額はカンマ・円記号・¥を除いた整数で返す
- 日付は必ずYYYY-MM-DD形式（和暦→西暦変換必須）

【和暦変換表】
令和7年=2025年、令和6年=2024年、令和5年=2023年、令和4年=2022年
令和3年=2021年、令和2年=2020年、令和元年=2019年

【読み取りのコツ】
- 手書き数字：1と7、6と0、3と8は文脈・桁数から判断
- 金額は「合計」「小計」「税込」「御請求金額」欄を最優先
- 複数の金額欄がある場合は最も大きい「税込合計」を合計金額とする
- 明細が複数行ある場合は全行を抽出する
- 仕入先は書類の発行元（右上・左上のスタンプ・印刷部分）から読む"""

    prompt_1st = f"""【書類の種類】{cat_hint}

この画像を注意深く観察し、以下のJSON形式のみで返してください。

{{
  "日付": "YYYY-MM-DD または null",
  "費目": "{category}",
  "明細": [
    {{
      "品名_作業内容": "文字列 または null",
      "数量": 数値またはnull,
      "単位": "個・本・袋・m・m²・m³・式・人・日・時間 など または null",
      "単価": 数値またはnull,
      "金額": 数値またはnull
    }}
  ],
  "合計金額": 数値またはnull,
  "消費税": 数値またはnull,
  "仕入先_外注先": "店名・会社名・個人名 または null",
  "読み取り信頼度": "高 または 中 または 低",
  "不明瞭箇所": ["読み取りに自信がない項目名をリストで"],
  "備考": "特記事項 または null"
}}"""

    # 1回目の読み取り
    # PDF・画像どちらも対応
    if is_pdf:
        file_content = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
    else:
        file_content = {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}}

    rotation_hint = """
- 画像が横向き・逆さ・斜めになっていても、書類の内容から正しい向きを判断して読み取ってください
- 書類の向きは文字の方向・金額の位置・日付の位置から判断してください
- 向きを補正した上で全ての情報を読み取ってください"""

    msg1 = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,
        system=system + ("\n- PDFの場合、全ページから情報を読み取り最初のページの書類を優先してください" if is_pdf else rotation_hint),
        messages=[{
            "role": "user",
            "content": [
                file_content,
                {"type": "text", "text": prompt_1st},
            ],
        }],
    )
    raw1   = re.sub(r"```json|```", "", msg1.content[0].text).strip()
    result = json.loads(raw1)

    # 信頼度が低・中の場合は2回目の読み取りで補完
    confidence = result.get("読み取り信頼度", "高")
    unclear    = result.get("不明瞭箇所", [])

    if confidence in ["低", "中"] and unclear:
        unclear_str = "・".join(unclear)
        prompt_2nd = f"""この画像をもう一度注意深く見てください。
特に「{unclear_str}」の部分が不明瞭でした。

前回の読み取り結果：
{json.dumps(result, ensure_ascii=False, indent=2)}

上記の不明瞭だった箇所に集中して再確認し、より正確な値に修正したJSONのみを返してください。
確信が持てない場合はnullのままにしてください。"""

        msg2 = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=2000,
            system=system,
            messages=[
                {
                    "role": "user",
                    "content": [
                        file_content,
                        {"type": "text", "text": prompt_1st},
                    ],
                },
                {"role": "assistant", "content": raw1},
                {"role": "user",     "content": prompt_2nd},
            ],
        )
        raw2 = re.sub(r"```json|```", "", msg2.content[0].text).strip()
        try:
            result2 = json.loads(raw2)
            # 2回目で改善された項目のみ上書き
            for key in ["日付", "合計金額", "消費税", "仕入先_外注先"]:
                if result2.get(key) is not None and result.get(key) is None:
                    result[key] = result2[key]
            if result2.get("明細") and len(result2["明細"]) >= len(result.get("明細") or []):
                result["明細"] = result2["明細"]
            if result2.get("読み取り信頼度") == "高":
                result["読み取り信頼度"] = "高"
            result["備考"] = f"2段階読み取り実施。{result.get('備考') or ''}"
        except Exception as e2:
            # 2回目の解析失敗時は1回目の結果をそのまま使用
            import logging
            logging.warning(f"2段階目の読み取り解析失敗: {e2}")

    # 不明瞭箇所リストは返却不要
    result.pop("不明瞭箇所", None)

    if not result.get("明細") and result.get("合計金額") is None:
        raise ValueError("readable_failed")
    return result

# =========================================================
# PCエージェント向けAPI（認証不要・内部用）
# =========================================================
@app.get("/api/queue")
async def get_queue():
    data = load()
    # PCエージェント向け：全キューと全プロジェクトを返す（PCエージェントは認証なしでアクセス）
    return {"queue": [r for r in data["queue"] if not r["done"]], "projects": data["projects"]}

@app.post("/api/queue/done")
async def mark_queue_done(request: Request):
    body = await request.json()
    ids  = set(body.get("ids", []))
    data = load()
    for r in data["queue"]:
        if r["id"] in ids:
            r["done"] = True
    save(data)
    return {"ok": True}

@app.post("/api/projects/done")
async def mark_project_done_from_pc(request: Request):
    body = await request.json()
    pid  = body.get("project_id")
    data = load()
    for p in data["projects"]:
        if p["id"] == pid:
            p["done"] = True
    save(data)
    return {"ok": True}

@app.get("/")
async def health():
    return {"status": "ok", "version": "3.0"}

# =========================================================
# 手打ち入力 → キューへ追加
# =========================================================
@app.post("/api/records/manual")
async def create_manual_record(request: Request, user_id: str = Depends(verify_token)):
    body          = await request.json()
    project_id    = body.get("project_id")
    category      = body.get("category")
    manual_result = body.get("manual_result")

    if not all([project_id, category, manual_result]):
        raise HTTPException(400, "project_id・category・manual_result は必須です")
    if category not in CATEGORIES:
        raise HTTPException(400, f"費目は {CATEGORIES} のいずれかを指定してください")

    data    = load()
    project = next((p for p in data["projects"] if p["id"] == project_id), None)
    if not project:
        raise HTTPException(404, "工事が見つかりません")

    record = {
        "id":           f"R{int(datetime.now().timestamp())}",
        "project_id":   project_id,
        "project_name": project["name"],
        "project_num":  project["num"],
        "category":     category,
        "ai_result":    manual_result,
        "queued_at":    datetime.now().isoformat(),
        "done":         False,
    }
    data["queue"].append(record)
    save(data)
    return {"record": record}


# =========================================================
# 一時保管用：画像のみAI読み取りしてキューには入れない
# 新規工事の場合は project_id なしで project_name 等を渡すと自動登録
# =========================================================
@app.post("/api/records/temp")
async def read_temp_record(request: Request, user_id: str = Depends(verify_token)):
    body        = await request.json()
    project_id  = body.get("project_id")
    image_b64   = body.get("image_b64")
    project_name= body.get("project_name")
    project_num = body.get("project_num", "")
    project_start=body.get("project_start", datetime.now().strftime("%Y-%m-%d"))
    project_person=body.get("project_person", "")

    if not image_b64:
        raise HTTPException(400, "image_b64 は必須です")

    # メディアタイプ判定（PDF対応）
    media_type = body.get("media_type", "image/jpeg")
    is_pdf = media_type == "application/pdf"

    data = load()

    # project_idがない場合は新規登録
    if not project_id:
        if not project_name:
            raise HTTPException(400, "project_id または project_name が必要です")
        # 同名の工事が既にあれば再利用（同一ユーザーのみ）
        existing = next((p for p in data["projects"] if p["name"] == project_name and not p.get("done") and p.get("owner") == user_id), None)
        if existing:
            project_id = existing["id"]
            project = existing
        else:
            project_id = f"P{int(datetime.now().timestamp())}"
            num = project_num or f"{datetime.now().year}-{len(data['projects'])+1:03d}"
            project = {
                "id": project_id, "name": project_name, "num": num,
                "start": project_start, "person": project_person,
                "location": body.get("project_location", ""),
                "contract_amount":     body.get("project_contract", ""),
                "contract_amount_ex":  body.get("project_contract_ex", ""),
                "contract_amount_tax": body.get("project_contract_tax", ""),
                "done": False, "owner": user_id
            }
            data["projects"].append(project)
            save(data)
    else:
        project = next((p for p in data["projects"] if p["id"] == project_id and p.get("owner") == user_id), None)
        if not project:
            raise HTTPException(404, "工事が見つかりません")

    try:
        result = await ai_read(base64.b64decode(image_b64), "不明", media_type=body.get("media_type", "image/jpeg"))
    except ValueError as e:
        raise HTTPException(422, detail={
            "message": "画像を読み取れませんでした。明るく・文字がはっきり写った写真で再度お試しください。",
            "error_type": "unreadable",
            "detail": str(e),
        })
    except json.JSONDecodeError as e:
        raise HTTPException(422, detail={
            "message": "AIの応答を解析できませんでした。再度お試しください。",
            "error_type": "json_parse_error",
            "detail": str(e),
        })
    except anthropic.APIError as e:
        raise HTTPException(502, detail={
            "message": f"Claude APIエラーが発生しました。しばらく待ってから再試行してください。",
            "error_type": "api_error",
            "detail": str(e),
        })
    except Exception as e:
        raise HTTPException(500, detail={
            "message": "AI読み取り中に予期しないエラーが発生しました。",
            "error_type": "unknown_error",
            "detail": str(e),
        })
    return {"ai_result": result, "project_id": project_id, "project": project}

# =========================================================
# 一時保管からエクスポート → Excelをブラウザにダウンロード
# =========================================================

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def build_excel(project: dict, category: str, records: list) -> bytes:
    """Excelファイルをメモリ上で生成してバイトとして返す"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category
    color = CAT_COLORS.get(category, "333333")

    # タイトル行
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{project['name']}　{category}台帳"
    ws["A1"].font      = Font(name="游ゴシック", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=color)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 工事情報行
    ws["A2"] = f"工事番号：{project.get('num','')}　開始日：{project.get('start','')}　請負金額：{project.get('contract_amount','')}円　工事場所：{project.get('location','')}"
    ws["A2"].font = Font(name="游ゴシック", size=10, color="666666")

    # ヘッダー行（備考は自由記載のためAIデータは入れない）
    headers = ["日付","費目","品名・作業内容","数量","単位","単価","金額","仕入先・外注先","備考（自由記載）"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[3].height = 22

    # 列幅
    for i, w in enumerate([12, 10, 30, 8, 6, 10, 12, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # データ行
    row = 4
    for ai in records:
        items = ai.get("明細", []) or [{"品名_作業内容": "", "数量": None, "単位": None, "単価": None, "金額": ai.get("合計金額", 0)}]
        for item in items:
            row_data = [
                ai.get("日付", ""), category,
                item.get("品名_作業内容", ""),
                item.get("数量"), item.get("単位"), item.get("単価"),
                item.get("金額") or 0,
                ai.get("仕入先_外注先", ""), "",  # 備考は自由記載のため空欄
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font   = Font(name="游ゴシック", size=10)
                c.border = _border()
                if col == 7 and val:
                    c.number_format = "#,##0"
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@app.post("/api/records/export")
async def export_records(request: Request, user_id: str = Depends(verify_token)):
    body       = await request.json()
    project_id = body.get("project_id")
    records    = body.get("records", [])   # [{category, ai_result, record_id}, ...]

    if not project_id or not records:
        raise HTTPException(400, "project_id・records は必須です")

    data    = load()
    project = next((p for p in data["projects"] if p["id"] == project_id and p.get("owner") == user_id), None)
    if not project:
        raise HTTPException(404, "工事が見つかりません")

    if "excel_cache"     not in data: data["excel_cache"]     = {}
    if "exported_records" not in data: data["exported_records"] = {}
    if "row_map"         not in data: data["row_map"]         = {}

    cache_key   = f"{user_id}_{project_id}"
    exported_ids = set(data["exported_records"].get(cache_key, []))
    row_map      = data["row_map"].get(cache_key, {})  # {record_id: {sheet, rows: [row_num,...]}}

    safe_name = project["name"].replace("/", "／")
    filename  = f"工事台帳_{safe_name}.xlsx"

    # Excelを読み込み or 新規作成
    if cache_key in data["excel_cache"]:
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data["excel_cache"][cache_key])))
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_sheet1(wb.create_sheet("工事台帳", 0), project)
        _build_sheet2(wb.create_sheet("集計", 1), project)
        for cat in CATEGORIES:
            _setup_cat_sheet(wb.create_sheet(cat), project, cat)

    added = 0
    overwritten = 0

    for r in records:
        cat       = r.get("category")
        record_id = r.get("record_id")
        ai        = r.get("ai_result", {})
        if cat not in CATEGORIES:
            continue

        if cat not in wb.sheetnames:
            _setup_cat_sheet(wb.create_sheet(cat), project, cat)
        ws = wb[cat]

        items = ai.get("明細", []) or [{"品名_作業内容":"","数量":None,"単位":None,"単価":None,"金額":ai.get("合計金額",0)}]

        def write_rows(start_row, item_list):
            written = []
            for idx, item in enumerate(item_list):
                row_num = start_row + idx
                row_data = [ai.get("日付",""), cat, item.get("品名_作業内容",""),
                            item.get("数量"), item.get("単位"), item.get("単価"),
                            item.get("金額") or 0, ai.get("仕入先_外注先",""), ""]
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_num, column=col, value=val)
                    c.font   = Font(name="游ゴシック", size=10)
                    c.border = _border()
                    if col == 7 and val:
                        c.number_format = "#,##0"
                written.append(row_num)
            return written

        if record_id and record_id in row_map:
            # 既存行を上書き
            existing = row_map[record_id]
            start_row = existing["rows"][0] if existing.get("rows") else None
            if start_row:
                # 既存行をクリアして上書き
                for row_num in existing["rows"]:
                    for col in range(1, 10):
                        ws.cell(row=row_num, column=col, value="")
                written = write_rows(start_row, items)
                row_map[record_id] = {"sheet": cat, "rows": written}
                overwritten += 1
            else:
                # 行情報なし → 追記
                next_row = 4
                while ws.cell(row=next_row, column=1).value is not None:
                    next_row += 1
                written = write_rows(next_row, items)
                row_map[record_id] = {"sheet": cat, "rows": written}
                added += 1
        else:
            # 新規追記
            next_row = 4
            while ws.cell(row=next_row, column=1).value is not None:
                next_row += 1
            written = write_rows(next_row, items)
            if record_id:
                row_map[record_id] = {"sheet": cat, "rows": written}
                exported_ids.add(record_id)
            added += 1

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    data["excel_cache"][cache_key]      = base64.b64encode(excel_bytes).decode()
    data["exported_records"][cache_key] = list(exported_ids)
    data["row_map"][cache_key]          = row_map
    save(data)
    return {"ok": True, "excel_b64": data["excel_cache"][cache_key], "filename": filename,
            "added": added, "overwritten": overwritten}


def _build_sheet1(ws, project):
    """シート1：工事台帳（テンプレートフォーマット）A4縦"""
    from openpyxl.styles import Alignment as Aln
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True

    def s(row, col, val="", bold=False, size=9, h="left", v="center", bg=None, wrap=False, border=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="游ゴシック", size=size, bold=bold)
        c.alignment = Aln(horizontal=h, vertical=v, wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if border:
            sd = Side(style="thin")
            c.border = Border(left=sd, right=sd, top=sd, bottom=sd)
        return c

    def m(r1,c1,r2,c2):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    HDR = "DDDDDD"
    col_w = [5,5,5,5,14,5,12,5,5,5,12,5,12,5]
    for i,w in enumerate(col_w,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r,h in [(1,16),(2,18),(3,18),(4,18),(5,18),(6,18),(7,22),(8,32),(9,18),(10,18),(11,18),(12,18),(13,18)]:
        ws.row_dimensions[r].height = h

    # タイトル
    m(1,1,1,14); s(1,1,"工　事　台　帳",bold=True,size=14,h="center",border=False)

    # 工事番号・工事名・工事場所
    m(2,1,2,2); s(2,1,"工事
番号",h="center",bg=HDR,wrap=True,size=7)
    m(2,3,2,5); s(2,3,project.get("num",""),size=9)
    m(2,6,3,6); s(2,6,"工
事
名",h="center",bg=HDR,wrap=True,size=7)
    m(2,7,2,10); s(2,7,project.get("name",""),bold=True,size=10)
    m(2,11,2,11); s(2,11,"工事
場所",h="center",bg=HDR,wrap=True,size=7)
    m(2,12,2,14); s(2,12,project.get("location",""),size=9)

    m(3,1,3,2); s(3,1,"",bg=HDR)
    m(3,3,3,5); s(3,3,"")
    m(3,7,3,10); s(3,7,"")
    m(3,12,3,13); s(3,12,"TEL")
    s(3,14,"",bg=HDR,size=7)

    # 発注者
    m(4,1,6,2); s(4,1,"発
注
者",h="center",bg=HDR,wrap=True,size=8)
    m(4,3,4,4); s(4,3,"名称",h="center",bg=HDR,size=8)
    m(4,5,4,9); s(4,5,project.get("orderer",""),size=9)
    s(4,10,"担当者",h="center",bg=HDR,size=7)
    m(4,11,4,14); s(4,11,project.get("person",""),size=9)

    m(5,3,5,4); s(5,3,"住所",h="center",bg=HDR,size=8)
    m(5,5,5,14); s(5,5,"",size=9)

    m(6,3,6,4); s(6,3,"TEL",h="center",bg=HDR,size=8)
    m(6,5,6,8); s(6,5,"",size=9)
    s(6,9,"確認番号",h="center",bg=HDR,size=7)
    m(6,10,6,11); s(6,10,f"第　　号",size=8)
    s(6,12,"確認
年月日",h="center",bg=HDR,size=7,wrap=True)
    m(6,13,6,14); s(6,13,"",size=9)

    # 工期
    m(7,1,7,2); s(7,1,"契約
年月日",h="center",bg=HDR,wrap=True,size=7)
    m(7,3,7,4); s(7,3,project.get("start",""),size=8)
    m(7,5,7,5); s(7,5,"着工
年月日",h="center",bg=HDR,wrap=True,size=7)
    s(7,6,"予定
実行",h="center",bg=HDR,wrap=True,size=7)
    m(7,7,7,7); s(7,7,project.get("start",""),size=8)
    m(7,8,7,8); s(7,8,"竣工
年月日",h="center",bg=HDR,wrap=True,size=7)
    s(7,9,"予定
実行",h="center",bg=HDR,wrap=True,size=7)
    m(7,10,7,10); s(7,10,project.get("completion_date",""),size=8)
    s(7,11,"引渡
年月日",h="center",bg=HDR,wrap=True,size=7)
    s(7,12,"予定
実行",h="center",bg=HDR,wrap=True,size=7)
    m(7,13,7,14); s(7,13,project.get("completion_date",""),size=8)

    # 工事概要
    m(8,1,8,2); s(8,1,"工事
概要",h="center",bg=HDR,wrap=True,size=8)
    m(8,3,8,14); s(8,3,"",wrap=True)

    # 請負金額セクション
    m(9,1,12,2); s(9,1,"請
負
金
額",h="center",bg=HDR,wrap=True,size=8)
    for col,label in [(3,"月/日"),(4,"摘　要"),(5,""),(6,"金　額"),(7,"月/日"),(8,"摘　要"),(9,""),(10,"金　額"),(11,"月/日"),(12,"摘　要"),(13,""),(14,"金　額")]:
        s(9,col,label,h="center",bg=HDR,size=8)
    for r in range(10,13):
        ws.row_dimensions[r].height = 16
        for c in range(3,15):
            s(r,c,"",size=9)

    # 支払条件
    m(13,1,13,2); s(13,1,"支払
条件",h="center",bg=HDR,wrap=True,size=8)
    m(13,3,13,14); s(13,3,"",wrap=True)

    # 請求・受入セクション
    row = 14
    s(row,1,"月/日",h="center",bg=HDR,size=8)
    m(row,2,row,4); s(row,2,"請　　求",h="center",bg=HDR,size=8)
    m(row,5,row,6); s(row,5,"金　額",h="center",bg=HDR,size=8)
    s(row,7,"月/日",h="center",bg=HDR,size=8)
    m(row,8,row,10); s(row,8,"受　　入",h="center",bg=HDR,size=8)
    m(row,11,row,12); s(row,11,"金　額",h="center",bg=HDR,size=8)
    m(row,13,row,13); s(row,13,"未収入金額",h="center",bg=HDR,size=7)
    s(row,14,"備考",h="center",bg=HDR,size=8)

    for r in range(row+1, row+13):
        ws.row_dimensions[r].height = 15
        s(r,1,"／",h="center",size=9)
        m(r,2,r,4); s(r,2,"",size=9)
        m(r,5,r,6); s(r,5,"",size=9)
        s(r,7,"／",h="center",size=9)
        m(r,8,r,10); s(r,8,"",size=9)
        m(r,11,r,12); s(r,11,"",size=9)
        s(r,13,"",size=9)
        s(r,14,"",size=9)

    # 請負金額の内訳（下部）
    bot = row + 13
    for r in [bot,bot+1,bot+2,bot+3,bot+4]:
        ws.row_dimensions[r].height = 16

    m(bot,1,bot,4); s(bot,1,"労　働　保　険　番　号",h="center",bg=HDR,size=8)
    m(bot,5,bot,14); s(bot,5,"請　負　金　額　の　内　訳",h="center",bg=HDR,size=8)

    m(bot+1,1,bot+2,1); s(bot+1,1,"府県",h="center",bg=HDR,size=7)
    m(bot+1,2,bot+2,2); s(bot+1,2,"所掌
管轄",h="center",bg=HDR,wrap=True,size=7)
    m(bot+1,3,bot+2,3); s(bot+1,3,"基幹
番号",h="center",bg=HDR,wrap=True,size=7)
    m(bot+1,4,bot+2,4); s(bot+1,4,"枝番号",h="center",bg=HDR,size=7,wrap=True)
    m(bot+1,5,bot+2,6); s(bot+1,5,"請負代金の額",h="center",bg=HDR,size=7,wrap=True)
    m(bot+1,7,bot+2,9); s(bot+1,7,"請負代金に
加算する額",h="center",bg=HDR,size=7,wrap=True)
    m(bot+1,10,bot+2,11); s(bot+1,10,"請負代金から
控除する額",h="center",bg=HDR,size=7,wrap=True)
    m(bot+1,12,bot+2,13); s(bot+1,12,"請負金額",h="center",bg=HDR,size=8,wrap=True)
    m(bot+1,14,bot+2,14); s(bot+1,14,"労務
費率",h="center",bg=HDR,wrap=True,size=7)

    for c in [1,2,3,4]:
        m(bot+3,c,bot+4,c); s(bot+3,c,"",size=9)
    m(bot+3,5,bot+3,6); s(bot+3,5,"",size=9)
    m(bot+3,7,bot+3,9); s(bot+3,7,"",size=9)
    m(bot+3,10,bot+3,11); s(bot+3,10,"",size=9)

    try:
        amt = int(str(project.get("contract_amount","0")).replace(",","").replace("円",""))
    except:
        amt = 0
    try:
        amt_ex = int(str(project.get("contract_amount_ex","0")).replace(",","").replace("円",""))
    except:
        amt_ex = 0
    try:
        tax = int(str(project.get("contract_amount_tax","0")).replace(",","").replace("円",""))
    except:
        tax = 0

    # 請負金額（税込）
    m(bot+2,12,bot+2,13)
    s(bot+2,12, amt if amt else "", h="right", size=10, bold=True)
    ws.cell(bot+2,12).number_format = "#,##0"
    ws.cell(bot+2,12).fill = PatternFill("solid", fgColor="FFF9C4")

    # 消費税を除く額
    s(bot+3,12, "消費税を除く額", h="right", size=7, bg="FFF9C4")
    s(bot+3,13, amt_ex if amt_ex else "", h="right", size=9)
    ws.cell(bot+3,13).number_format = "#,##0"

    # 消費税額
    s(bot+4,12, "消費税額", h="right", size=7, bg="FFF9C4")
    s(bot+4,13, tax if tax else "", h="right", size=9)
    ws.cell(bot+4,13).number_format = "#,##0"

    m(bot+3,14,bot+4,14); s(bot+3,14,"",size=9)
    ws.print_area = f"A1:N{bot+4}"


def _build_sheet2(ws, project):
    """シート2：集計シート（材料費・人件費・外注費・経費）A4縦"""
    from openpyxl.styles import Alignment as Aln
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"

    HDR_BLUE = "1565C0"
    CAT_C = {"材料費":"1565C0","人件費":"2E7D32","外注費":"6A1B9A","経費":"E65100"}
    sd = Side(style="thin")
    bdr = Border(left=sd,right=sd,top=sd,bottom=sd)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 14

    ws.merge_cells("A1:C1")
    ws["A1"] = f"集　計　表　　{project['name']}"
    ws["A1"].font = Font(name="游ゴシック", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="333333")
    ws["A1"].alignment = Aln(horizontal="center", vertical="center")
    ws["A1"].border = bdr

    ws["A2"] = f"工事番号：{project.get('num','')}　開始日：{project.get('start','')}　完成：{project.get('completion_date','')}"
    ws["A2"].font = Font(name="游ゴシック", size=8, color="666666")
    ws.merge_cells("A2:C2")

    # ヘッダー
    ws.row_dimensions[3].height = 18
    for col,label in [(1,"費　目"),(2,"合計金額（円）"),(3,"備　考")]:
        c = ws.cell(row=3,column=col,value=label)
        c.font = Font(name="游ゴシック",bold=True,size=10,color="1A1A1A")
        c.fill = PatternFill("solid",fgColor="D9D9D9")
        c.alignment = Aln(horizontal="center",vertical="center")
        c.border = bdr

    # 費目行
    row = 4
    for cat in CATEGORIES:
        ws.row_dimensions[row].height = 22
        c1 = ws.cell(row=row,column=1,value=cat)
        c1.font = Font(name="游ゴシック",bold=True,size=10,color="1A1A1A")
        c1.fill = PatternFill("solid",fgColor="EEEEEE")
        c1.alignment = Aln(horizontal="center",vertical="center")
        c1.border = bdr
        # 費目シートが存在する場合のみSUMIF
        c2 = ws.cell(row=row,column=2,value=f"=IFERROR(SUMPRODUCT(('{cat}'!B4:B2000="{cat}")*'{cat}'!G4:G2000),0)")
        c2.font = Font(name="游ゴシック",size=10)
        c2.number_format = "#,##0"
        c2.alignment = Aln(horizontal="right",vertical="center")
        c2.border = bdr
        c3 = ws.cell(row=row,column=3,value="")
        c3.border = bdr
        row += 1

    # 合計
    ws.row_dimensions[row].height = 24
    ct = ws.cell(row=row,column=1,value="合　計")
    ct.font = Font(name="游ゴシック",bold=True,size=11,color="FFFFFF")
    ct.fill = PatternFill("solid",fgColor="222222")
    ct.alignment = Aln(horizontal="center",vertical="center")
    ct.border = bdr
    cv = ws.cell(row=row,column=2,value=f"=SUM(B4:B{row-1})")
    cv.font = Font(name="游ゴシック",bold=True,size=11)
    cv.number_format = "#,##0"
    cv.alignment = Aln(horizontal="right",vertical="center")
    cv.border = bdr
    c3 = ws.cell(row=row,column=3,value="")
    c3.border = bdr

    # 請負金額との対比
    row += 2
    ws.row_dimensions[row].height = 18
    try:
        amt = int(str(project.get("contract_amount","0")).replace(",","").replace("円",""))
    except:
        amt = 0
    try:
        amt_ex = int(str(project.get("contract_amount_ex","0")).replace(",","").replace("円",""))
    except:
        amt_ex = 0
    try:
        tax = int(str(project.get("contract_amount_tax","0")).replace(",","").replace("円",""))
    except:
        tax = 0

    for r_offset, label, val in [(0,"請負金額（税込）",amt),(1,"消費税を除く額",amt_ex),(2,"消費税額",tax)]:
        ws.row_dimensions[row+r_offset].height = 18
        lc = ws.cell(row=row+r_offset,column=1,value=label)
        lc.font = Font(name="游ゴシック",size=9,color="555555")
        lc.alignment = Aln(horizontal="right",vertical="center")
        lc.border = bdr
        lc.fill = PatternFill("solid",fgColor="FFF9C4")
        vc = ws.cell(row=row+r_offset,column=2,value=val if val else "")
        vc.font = Font(name="游ゴシック",size=9)
        vc.number_format = "#,##0"
        vc.alignment = Aln(horizontal="right",vertical="center")
        vc.border = bdr
        ec = ws.cell(row=row+r_offset,column=3,value="")
        ec.border = bdr


def _setup_cat_sheet(ws, project, cat):
    """費目シートの初期設定（A4縦）"""
    from openpyxl.styles import Alignment as Aln
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{project['name']}　{cat}台帳"
    ws["A1"].font      = Font(name="游ゴシック", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="333333")
    ws["A1"].alignment = Aln(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws["A2"] = f"工事番号：{project.get('num','')}　担当：{project.get('person','')}　請負金額：{project.get('contract_amount','')}円　場所：{project.get('location','')}"
    ws["A2"].font = Font(name="游ゴシック", size=10, color="666666")
    ws["A2"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws.merge_cells("A2:I2")
    headers = ["日付","費目","品名・作業内容","数量","単位","単価","金額","仕入先・外注先","備考（自由記載）"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="555555")
        c.alignment = Aln(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[3].height = 22
    for i, w in enumerate([12, 10, 30, 8, 6, 10, 12, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return ws







# =========================================================
# Excelキャッシュリセット（工事単位）
# =========================================================
@app.delete("/api/records/export/{project_id}")
async def reset_export_cache(project_id: str, user_id: str = Depends(verify_token)):
    data = load()
    cache_key = f"{user_id}_{project_id}"
    removed = False
    if "excel_cache" in data and cache_key in data["excel_cache"]:
        del data["excel_cache"][cache_key]
        removed = True
    if "exported_records" in data and cache_key in data["exported_records"]:
        del data["exported_records"][cache_key]
    if "row_map" in data and cache_key in data["row_map"]:
        del data["row_map"][cache_key]
    save(data)
    return {"ok": True, "reset": removed}

# =========================================================
# 工事経歴書 Excelエクスポート
# =========================================================
@app.post("/api/career/export")
async def export_career(request: Request, user_id: str = Depends(verify_token)):
    body     = await request.json()
    year     = body.get("year")          # 対象年度（例：2024）
    company  = body.get("company", "")   # 申請者名
    permit   = body.get("permit", "")    # 許可番号
    work_type= body.get("work_type", "") # 建設工事の種類

    data = load()
    projects = [p for p in data["projects"] if p.get("owner") == user_id]

    # 年度フィルタ
    if year:
        def in_year(p):
            start = p.get("start","")
            comp  = p.get("completion_date","")
            try:
                s_year = int(start[:4]) if start else 0
                c_year = int(comp[:4])  if comp  else 0
                return s_year == int(year) or c_year == int(year)
            except:
                return False
        projects = [p for p in projects if in_year(p)]

    # Excelを生成
    excel_bytes = build_career_excel(projects, year, company, permit, work_type)
    excel_b64   = base64.b64encode(excel_bytes).decode()
    filename    = f"工事経歴書_{year or '全期間'}.xlsx"
    return {"ok": True, "excel_b64": excel_b64, "filename": filename}


def build_career_excel(projects, year, company, permit, work_type):
    from openpyxl.styles import Alignment as Aln
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year or '全期間'}年度"

    # ── カラー・スタイル定義 ──
    def b(row, col, val="", bold=False, size=9, align_h="center", align_v="center",
          bg=None, border_all=True, wrap=True, merge_to=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="游ゴシック", size=size, bold=bold)
        c.alignment = Aln(horizontal=align_h, vertical=align_v, wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if border_all:
            side = Side(style="thin")
            c.border = Border(left=side, right=side, top=side, bottom=side)
        return c

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    # ── 列幅設定 ──
    col_widths = [18, 6, 4, 22, 12, 8, 8, 12, 10, 4, 4, 4, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── タイトル行 ──
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 16

    ws["A1"] = "様式第二号（第二条、第十三条の二、第十三条の三、第十九条の八関係）"
    ws["A1"].font = Font(name="游ゴシック", size=7)
    ws["N1"] = "（用紙A４）"
    ws["N1"].font = Font(name="游ゴシック", size=7)
    ws["N1"].alignment = Aln(horizontal="right")

    # タイトル中央
    merge(2, 1, 2, 8)
    ws["A2"] = "工　　　事　　　経　　　歴　　　書"
    ws["A2"].font = Font(name="游ゴシック", size=14, bold=True)
    ws["A2"].alignment = Aln(horizontal="center", vertical="center")

    # 右側：決算期間・許可番号・申請者
    period_start = f"令和{int(str(year)[2:])if year else ''}年01月01日" if year else ""
    period_end   = f"令和{int(str(year)[2:])if year else ''}年12月31日" if year else ""
    ws["I2"] = f"決算期間　{period_start}〜{period_end}
許可番号　{permit}
申請者　　{company}"
    ws["I2"].font = Font(name="游ゴシック", size=8)
    ws["I2"].alignment = Aln(horizontal="left", vertical="center", wrap_text=True)
    merge(2, 9, 2, 14)

    # 工事種類行
    merge(3, 1, 3, 8)
    ws["A3"] = f"（建設工事の種類）　　{work_type}　　工事（税込・税抜）"
    ws["A3"].font = Font(name="游ゴシック", size=9)

    # ── ヘッダー行（4〜6行） ──
    HDR_BG = "D9D9D9"
    for r in [4, 5, 6]:
        ws.row_dimensions[r].height = 20

    # 注文者
    merge(4, 1, 6, 1); b(4, 1, "注
文
者", bold=True, bg=HDR_BG, wrap=True)
    # 元請又は下請の別
    merge(4, 2, 6, 2); b(4, 2, "元請
又は
下請
の別", bold=True, bg=HDR_BG, size=7, wrap=True)
    # JVの別
    merge(4, 3, 6, 3); b(4, 3, "J
V
の
別", bold=True, bg=HDR_BG, size=7, wrap=True)
    # 工事名
    merge(4, 4, 6, 4); b(4, 4, "工　　　事　　　名", bold=True, bg=HDR_BG)
    # 工事場所
    merge(4, 5, 6, 5); b(4, 5, "工事現場のある
都道府県及び市
区町村名", bold=True, bg=HDR_BG, size=7, wrap=True)
    # 配置技術者
    merge(4, 6, 4, 7); b(4, 6, "配　置　技　術　者", bold=True, bg=HDR_BG)
    b(5, 6, "氏　名", bold=True, bg=HDR_BG)
    b(5, 7, "主任技術者又は監理技術者
の別（該当箇所にレ印を記載）", bold=True, bg=HDR_BG, size=6, wrap=True)
    b(6, 6, "主任技術者", bold=True, bg=HDR_BG, size=7)
    b(6, 7, "監理技術者", bold=True, bg=HDR_BG, size=7)
    # 請負代金の額
    merge(4, 8, 4, 11); b(4, 8, "請　負　代　金　の　額", bold=True, bg=HDR_BG)
    merge(5, 8, 6, 8); b(5, 8, "（千円）", bold=True, bg=HDR_BG, size=7)
    merge(5, 9, 5, 11); b(5, 9, "うち、
・ＰＣ
・法面処理
・鋼橋上部", bold=True, bg=HDR_BG, size=6, wrap=True)
    b(6, 9, "ＰＣ", bold=True, bg=HDR_BG, size=7)
    b(6, 10, "法面
処理", bold=True, bg=HDR_BG, size=7, wrap=True)
    b(6, 11, "鋼橋
上部", bold=True, bg=HDR_BG, size=7, wrap=True)
    # 工期
    merge(4, 12, 4, 14); b(4, 12, "工　　　　　期", bold=True, bg=HDR_BG)
    merge(5, 12, 6, 12); b(5, 12, "着工年月", bold=True, bg=HDR_BG, size=7)
    merge(5, 13, 6, 14); b(5, 13, "完成又は
完成予定年月", bold=True, bg=HDR_BG, size=7, wrap=True)

    # ── データ行（7行目〜） ──
    ROW_START = 7
    MAX_ROWS  = 15

    def fmt_date(d):
        if not d: return ""
        try:
            from datetime import datetime as dt
            parsed = dt.strptime(d[:7], "%Y-%m")
            reiwa = parsed.year - 2018
            return f"令和{reiwa}年{parsed.month}月"
        except:
            return d

    for i in range(MAX_ROWS):
        r = ROW_START + i
        ws.row_dimensions[r].height = 18
        if i < len(projects):
            p = projects[i]
            amt = p.get("contract_amount","")
            try: amt_int = int(str(amt).replace(",","").replace("円","")) // 1000
            except: amt_int = ""

            b(r, 1,  p.get("orderer",""),        align_h="left")
            b(r, 2,  p.get("jv_type","元請"),    size=8)
            b(r, 3,  "",                          size=8)
            b(r, 4,  p.get("name",""),            align_h="left")
            b(r, 5,  p.get("location",""),        size=8, align_h="left")
            b(r, 6,  p.get("engineer_name",""),   size=8)
            b(r, 7,  "レ" if p.get("engineer_chief") else "", size=8)
            # 監理技術者は別列に
            b(r, 8,  f"{amt_int}" if amt_int else "", align_h="right")
            ws.cell(r, 8).number_format = '#,##0'
            b(r, 9,  "レ" if p.get("has_pc") else "")
            b(r, 10, "レ" if p.get("has_surface") else "")
            b(r, 11, "レ" if p.get("has_steel") else "")
            b(r, 12, fmt_date(p.get("start","")), size=8)
            b(r, 13, fmt_date(p.get("completion_date","")), size=8)
            merge(r, 13, r, 14)
        else:
            for col in range(1, 15):
                b(r, col, "")
            # 千円表示
            b(r, 8, "", align_h="right")

    # ── 小計・合計行 ──
    sum_row = ROW_START + MAX_ROWS
    ws.row_dimensions[sum_row].height = 22
    ws.row_dimensions[sum_row+2].height = 22

    total_amt = sum(
        int(str(p.get("contract_amount","0")).replace(",","").replace("円","")) // 1000
        for p in projects if p.get("contract_amount")
    ) if projects else 0
    primary = sum(1 for p in projects if p.get("jv_type","元請") == "元請")
    primary_amt = sum(
        int(str(p.get("contract_amount","0")).replace(",","").replace("円","")) // 1000
        for p in projects if p.get("jv_type","元請") == "元請" and p.get("contract_amount")
    )

    merge(sum_row, 1, sum_row, 6); b(sum_row, 1, "小　計", bold=True, bg=HDR_BG)
    b(sum_row, 7, f"{len(projects)}件", bold=True, bg=HDR_BG)
    b(sum_row, 8, total_amt if total_amt else "", bold=True, bg=HDR_BG, align_h="right")
    ws.cell(sum_row, 8).number_format = '#,##0'
    for col in range(9, 15):
        b(sum_row, col, "", bg=HDR_BG)

    # うち元請工事
    merge(sum_row, 12, sum_row, 13)
    ws.cell(sum_row, 12).value = "うち元請工事"
    ws.cell(sum_row, 12).font = Font(name="游ゴシック", size=7, bold=True)
    ws.cell(sum_row, 12).alignment = Aln(horizontal="center")
    b(sum_row, 14, primary_amt if primary_amt else "", bold=True, align_h="right")
    ws.cell(sum_row, 14).number_format = '#,##0'

    # 合計（全体）
    merge(sum_row+2, 1, sum_row+2, 6); b(sum_row+2, 1, "合　計", bold=True, bg=HDR_BG)
    b(sum_row+2, 7, f"{len(projects)}件", bold=True, bg=HDR_BG)
    b(sum_row+2, 8, total_amt if total_amt else "", bold=True, bg=HDR_BG, align_h="right")
    ws.cell(sum_row+2, 8).number_format = '#,##0'
    for col in range(9, 15):
        b(sum_row+2, col, "", bg=HDR_BG)

    merge(sum_row+2, 12, sum_row+2, 13)
    ws.cell(sum_row+2, 12).value = "うち元請工事"
    ws.cell(sum_row+2, 12).font = Font(name="游ゴシック", size=7, bold=True)
    ws.cell(sum_row+2, 12).alignment = Aln(horizontal="center")
    b(sum_row+2, 14, primary_amt if primary_amt else "", bold=True, align_h="right")
    ws.cell(sum_row+2, 14).number_format = '#,##0'

    ws.print_area = f"A1:N{sum_row+3}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
